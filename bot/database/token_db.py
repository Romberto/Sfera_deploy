import datetime
import io
import csv
from decimal import Decimal
from pprint import pprint

import json
from openpyxl.workbook import Workbook

from database.config_db import connect
from collections import namedtuple
import asyncio

Token = namedtuple('Token', [
    'id', 'active', 'body', 'date_create', 'date_activate', 'driver_activate', 'excursion_name', 'excursion_price'
])


async def get_token(token_text: str):
    conn = connect()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT tokens_tokenexmodel.id, tokens_tokenexmodel.active,"
                           "tokens_tokenexmodel.body,tokens_tokenexmodel.date_create,"
                           "tokens_tokenexmodel.date_activate,tokens_tokenexmodel.driver_activate_id"
                           ", excursions_excursionmodel.name, excursions_excursionmodel.price "
                           "FROM tokens_tokenexmodel RIGHT JOIN excursions_excursionmodel "
                           "ON tokens_tokenexmodel.excursion_id = excursions_excursionmodel.id WHERE "
                           "body='" + token_text + "' AND active = 'True'")
            record = cursor.fetchone()
            token_db = Token(
                id=record[0],
                active=record[1],
                body=record[2],
                date_create=record[3],
                date_activate=record[4],
                driver_activate=record[5],
                excursion_name=record[6],
                excursion_price=record[7]
            )
            cursor.close()
            return token_db
        except Exception as _err:
            print('token_db.py => get_token', _err)
            return False
        finally:
            conn.close()
    else:
        print("проблемы с подключением к базе")


async def token_finaly(token: str, driver):
    conn = connect()
    if conn:
        today = datetime.datetime.now()
        try:
            cursor = conn.cursor()
            cursor.execute("UPDATE tokens_tokenexmodel SET active='False', date_activate='" + today.strftime(
                '%Y-%m-%d %H:%M:%S') + f"', driver_activate_id='{driver}' WHERE body='" + token + "';")
            conn.commit()
            cursor.close()
            return True
        except Exception as _err:
            print('token_db.py => token_finaly', _err)
            return False
        finally:
            conn.close()
    else:
        print("проблемы с подключением к базе")


async def check_driver_admin(id: str):
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM drivers_drivermodel WHERE bot_user_id='" + id + "';")
        record = cursor.fetchone()
        driver = {'id': record[0], 'admin_bool': record[9], 'login': record[4], 'drivers_group': record[10], 'record': record}
        conn.close()
        return driver
    except TypeError as _err:
        return False


async def personal_sales(driver_activate_id):
    """
    личные продажи по принимает id drivers
    запрос к базе
    возвращает список
    """
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT tokens_tokenexmodel.id, tokens_tokenexmodel.active,"
            "tokens_tokenexmodel.body,"
            "tokens_tokenexmodel.date_activate,tokens_tokenexmodel.driver_activate_id"
            ", excursions_excursionmodel.name, excursions_excursionmodel.price "
            "FROM tokens_tokenexmodel RIGHT JOIN excursions_excursionmodel "
            "ON tokens_tokenexmodel.excursion_id = excursions_excursionmodel.id WHERE "
            "driver_activate_id='" + str(driver_activate_id) + "' AND active = 'False'"
                                                               f"AND DATE(date_activate) = DATE(NOW())"
        )
        record = cursor.fetchall()

        return record
    except TypeError as _err:
        print(_err)
        return False


async def personal_sales_to_str(obj):
    """
    принимает картеж и приводит ве эленты кортежа к строке
    """
    result_list = ()
    for item in obj:
        if type(item) == datetime.datetime:
            date = item.strftime('%d-%m-%Y %H:%M')
            result_list = result_list + (date,)
        else:
            result_list = result_list + (str(item),)
    # print(result_list)
    return result_list


async def make_report(id_driver):
    """
    принимае id водителя и генерирует список кортежей из базы для отчёта
    возвращает лист с кортежами
    """
    query = await personal_sales(id_driver)
    if query:
        query_result_list = []
        for item in query:
            item_set = await personal_sales_to_str(item)
            query_result_list.append(item_set)
        return query_result_list


async def make_lxml(id_driver, login_driver):
    """
    формирует и возвращает xlsx в виде байтов
    """
    wb = Workbook()
    ws = wb.active
    ws['A1'].value = 'id экскурии'
    ws['B1'].value = 'токен active'
    ws['C1'].value = 'код'
    ws['D1'].value = 'дата активации'
    ws['E1'].value = 'id водителя'
    ws['F1'].value = 'название маршрута'
    ws['G1'].value = 'цена маршрута'
    query = await make_report(id_driver)
    if query:
        token = 0  # общее количество проданных токенов
        summ = 0  # общая сумма
        for row in query:
            token += 1
            summ += float(row[6])
            ws.append(row)
        end_row = ('количество токенов:', f'{token}', '', 'сумма :', f'{summ}')
        ws.append((' ', ' '))
        ws.append(end_row)
        buf = io.BytesIO()
        buf.name = f'отчёт_{login_driver}.xlsx'
        buf.encoding = "utf-8"
        wb.save(buf)
        buf.seek(0)
        mes = f"* Сегодня {datetime.datetime.today().strftime('%d-%m-%Y')}\n* Логин - : {login_driver}\n* Kоличество токенов - {token}\n* Сумма - : {summ}"
        return (buf, mes)
    else:
        return False


async def all_sales_query(id_group):
    """
    принимает id группы водителей
    """
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT drivers_drivermodel.driver_last_name, drivers_drivermodel.login, drivers_drivermodel.drivers_group_id, "
            "excursions_excursionmodel.price "
            " FROM drivers_drivermodel "
            "RIGHT JOIN tokens_tokenexmodel "
            "ON drivers_drivermodel.id = tokens_tokenexmodel.driver_activate_id "
            "RIGHT JOIN excursions_excursionmodel "
            "ON tokens_tokenexmodel.excursion_id = excursions_excursionmodel.id "
            f"WHERE drivers_drivermodel.drivers_group_id={id_group} "
            f"AND active = 'False' "
            f"AND DATE(tokens_tokenexmodel.date_activate)=DATE(NOW())"
        )
        record = cursor.fetchall()
        return record
    except TypeError as _err:
        print(_err)
        return False


async def get_json(id_group):
    """
    принимает список query_set от all_sales_query
    возвращает json
    """
    query = await all_sales_query(id_group)
    if query:
        data = {
            'logins': set(),
            'tokens': {},
            'drivers_summ': {},
            'last_name': {}
        }
        for item in query:
            data['logins'].add(item[1])
            try:
                if data['drivers_summ'][item[1]]:
                    data['drivers_summ'][item[1]] += float(item[3])
                    data['tokens'][item[1]] += 1
                    data['last_name'][item[1]] = item[0]

            except KeyError:
                data['drivers_summ'].update({item[1]: float(item[3])})
                data['tokens'].update({item[1]: 1})
                data['last_name'].update({item[1]: item[0]})

        return data
    else:
        return False


query = [('test', 'ted', 3, Decimal('5000.00')), ('test', 'ted', 3, Decimal('7000.00')),
         ('test2', 'erl', 3, Decimal('7000.00'))]
data = {'drivers_summ': {'erl': 7000.0, 'ted': 12000.0},
        'last_name': {'erl': 'test2', 'ted': 'test'},
        'logins': {'erl', 'ted'},
        'tokens': {'erl': 1, 'ted': 2}}


async def answer_all_sales(id_group):
    data = await get_json(id_group)
    if data:
        count_driver = len(data['logins'])
        result_summ = 0
        result_token = 0

        data_list = []
        for item in data['logins']:
            summ  = data['drivers_summ'][item]
            last_name = data['last_name'][item]
            tokens = data['tokens'][item]
            result_summ += float(summ)
            result_token += int(tokens)
            data_set = (item, last_name, summ, tokens)
            data_list.append(data_set)
        wb = Workbook()
        ws = wb.active
        ws['A1'].value = 'логин'
        ws['B1'].value = 'фамилия'
        ws['C1'].value = 'сумма'
        ws['D1'].value = 'токены'
        for row in data_list:
            ws.append(row)
        row = (' ', " ")
        ws.append(row)
        row = (' ', 'Итого :', f'{result_summ}', f'{result_token}')
        ws.append(row)
        buf = io.BytesIO()
        buf.name = f'отчёт_{id_group}.xlsx'
        buf.encoding = "utf-8"
        wb.save(buf)
        buf.seek(0)
        mess = f"Сегодня {datetime.datetime.today().strftime('%d-%m-%Y')}\n" \
               f"* Кол-во водителей - {count_driver}\n" \
               f"* Кол-во токенов - {result_token}\n" \
               f"* Общая сумма - {result_summ} руб."
        return (buf, mess)
    else:
        return False

async def main():
    task1 = asyncio.create_task(
        answer_all_sales('3'))
    await task1


if __name__ == '__main__':
    asyncio.run(main())
